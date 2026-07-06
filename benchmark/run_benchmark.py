#!/usr/bin/env python3
"""
GeoBenchX Istanbul CBS Benchmark — Otomasyon Scripti
====================================================

GeoBenchX (Krechetova & Kochedykov, 2025) puanlama metodolojisini
İstanbul CBS senaryolarına uyarlayarak, mevcut n8n webhook altyapısını
köprü olarak kullanıp 7 LLM modelini otomatik test eder.

Kullanım:
    python benchmark/run_benchmark.py                # Tam benchmark
    python benchmark/run_benchmark.py --dry-run      # Bağlantı olmadan test
    python benchmark/run_benchmark.py --scenario TR_01  # Tek senaryo
    python benchmark/run_benchmark.py --delay 5      # 5 sn ara ile

Gereksinimler:
    pip install httpx openpyxl
    uvicorn main:app --reload --port 8050  (ayrı terminalde)
"""

import argparse
import asyncio
import json
import os
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import httpx

# ──────────────────────────────────────────────────────────────────────
# Sabitler
# ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
SCENARIOS_FILE = SCRIPT_DIR / "scenarios.json"
GROUND_TRUTH_FILE = SCRIPT_DIR / "ground_truth.json"
RESULTS_DIR = SCRIPT_DIR / "results"
EXCEL_FILE = PROJECT_DIR / "CBS_LLM_Dogruluk_ArcGIS_Karsilastirma.xlsx"

API_BASE_URL = os.environ.get("BENCHMARK_API_URL", "http://localhost:8050")
ANALYZE_ENDPOINT = f"{API_BASE_URL}/api/analyze"
HEALTH_TIMEOUT = 10.0
REQUEST_TIMEOUT = 120.0
DEFAULT_DELAY = 3.0

# 7 aktif model (kimi hariç, backend'de var ama frontend'de yok)
ACTIVE_MODELS = [
    {"key": "claude",   "name": "Claude Sonnet",  "color": "#d97757"},
    {"key": "gpt",      "name": "GPT-4o",         "color": "#19c37d"},
    {"key": "groq",     "name": "Groq (Llama 3)", "color": "#f55036"},
    {"key": "gemini",   "name": "Gemini",          "color": "#5b8def"},
    {"key": "deepseek", "name": "DeepSeek V3",    "color": "#3b82f6"},
    {"key": "qwen",     "name": "Qwen 2.5-72B",   "color": "#a855f7"},
    {"key": "glm",      "name": "GLM-4 Flash",    "color": "#14b8a6"},
]

MODEL_KEYS = [m["key"] for m in ACTIVE_MODELS]

# GeoBenchX tool eşleşme aliases — farklı isimlendirmeleri kabul eder
TOOL_ALIASES = {
    "Buffer":              ["buffer", "tampon", "st_buffer", "create_buffer"],
    "AttributeTable":      ["attributetable", "attribute_table", "select_by_attribute",
                            "list", "liste", "query", "filter"],
    "AttributeQuery":      ["attributequery", "attribute_query", "select_by_attribute",
                            "filter", "search", "arama"],
    "NearestSearch":       ["nearestsearch", "nearest_search", "near", "near_analysis",
                            "closest", "proximity", "en_yakin"],
    "PointDensity":        ["pointdensity", "point_density", "kernel_density",
                            "heatmap", "yogunluk", "density"],
    "Geocode":             ["geocode", "geocoding", "konum", "locate", "find_location"],
    "Intersection":        ["intersection", "intersect", "kesisim", "clip", "overlay"],
    "Routing":             ["routing", "route", "network_analysis", "network",
                            "rota", "navigation", "shortest_path"],
    "CoordinateTransform": ["coordinatetransform", "coordinate_transform", "project",
                            "transform", "crs", "projeksiyon", "donusum"],
}

# ──────────────────────────────────────────────────────────────────────
# Yardımcı fonksiyonlar
# ──────────────────────────────────────────────────────────────────────

class Colors:
    """ANSI renk kodları — terminalde renkli çıktı."""
    BOLD    = "\033[1m"
    RED     = "\033[91m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    BLUE    = "\033[94m"
    MAGENTA = "\033[95m"
    CYAN    = "\033[96m"
    DIM     = "\033[2m"
    RESET   = "\033[0m"


def cprint(msg: str, color: str = ""):
    """Renkli print."""
    print(f"{color}{msg}{Colors.RESET}")


def normalize_text(text: str) -> str:
    """Metin normalleştirme: küçük harf, Unicode normalize, boşluk trim."""
    if not text:
        return ""
    text = unicodedata.normalize("NFC", text.strip().lower())
    # Türkçe karakter dönüşümleri (karşılaştırma kolaylığı)
    replacements = {"ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c"}
    for tr_char, ascii_char in replacements.items():
        text = text.replace(tr_char, ascii_char)
    return text


def load_json(filepath: Path) -> dict:
    """JSON dosyasını yükle."""
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(data: dict, filepath: Path):
    """JSON dosyasını kaydet."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def timestamp_id() -> str:
    """Zaman damgası ID üret."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ──────────────────────────────────────────────────────────────────────
# 1. Health Check
# ──────────────────────────────────────────────────────────────────────

async def health_check(client: httpx.AsyncClient) -> bool:
    """
    Backend'in ayakta olup olmadığını kontrol eder.
    GET / endpoint'i 200 dönmeli.
    """
    cprint("\n🏥 Health Check — Backend bağlantısı kontrol ediliyor...", Colors.CYAN)
    try:
        resp = await client.get(API_BASE_URL, timeout=HEALTH_TIMEOUT)
        if resp.status_code == 200:
            cprint("   ✅ Backend aktif (HTTP 200)", Colors.GREEN)
            return True
        else:
            cprint(f"   ⚠️  Backend yanıt verdi ama HTTP {resp.status_code}", Colors.YELLOW)
            return True  # Yanıt var, devam edebiliriz
    except httpx.ConnectError:
        cprint(f"   ❌ Backend'e bağlanılamıyor: {API_BASE_URL}", Colors.RED)
        cprint("   ℹ️  'uvicorn main:app --reload --port 8050' komutunu çalıştırdığınızdan emin olun.", Colors.DIM)
        return False
    except httpx.TimeoutException:
        cprint(f"   ❌ Backend zaman aşımına uğradı ({HEALTH_TIMEOUT}s)", Colors.RED)
        return False
    except Exception as exc:
        cprint(f"   ❌ Beklenmeyen hata: {exc}", Colors.RED)
        return False


async def n8n_probe(client: httpx.AsyncClient) -> bool:
    """
    Basit bir test sorgusuyla n8n webhook'larının yanıt verip vermediğini kontrol eder.
    Tam benchmark başlamadan önce hızlı bir prob gönderir.
    """
    cprint("🔗 n8n Webhook Probe — Bir test sorgusu gönderiliyor...", Colors.CYAN)
    try:
        resp = await client.post(
            ANALYZE_ENDPOINT,
            json={"message": "test"},
            timeout=REQUEST_TIMEOUT,
        )
        if resp.status_code == 200:
            data = resp.json()
            active_count = sum(
                1 for key in MODEL_KEYS
                if key in data and data[key].get("status") != "error"
            )
            error_count = sum(
                1 for key in MODEL_KEYS
                if key in data and data[key].get("status") == "error"
            )
            total = active_count + error_count
            cprint(f"   ✅ n8n yanıt verdi: {active_count}/{total} model aktif", Colors.GREEN)
            if error_count > 0:
                error_models = [
                    key for key in MODEL_KEYS
                    if key in data and data[key].get("status") == "error"
                ]
                cprint(f"   ⚠️  Hata veren modeller: {', '.join(error_models)}", Colors.YELLOW)
            return True
        else:
            cprint(f"   ⚠️  Backend HTTP {resp.status_code} döndü", Colors.YELLOW)
            return False
    except Exception as exc:
        cprint(f"   ❌ n8n probe başarısız: {exc}", Colors.RED)
        return False


# ──────────────────────────────────────────────────────────────────────
# 2. Senaryo Yükleme
# ──────────────────────────────────────────────────────────────────────

def load_scenarios(scenario_filter: str = None) -> list:
    """Senaryoları JSON'dan yükle, isteğe bağlı filtrele."""
    data = load_json(SCENARIOS_FILE)
    scenarios = data.get("scenarios", [])

    if scenario_filter:
        scenarios = [s for s in scenarios if s["id"] == scenario_filter]
        if not scenarios:
            cprint(f"❌ Senaryo '{scenario_filter}' bulunamadı!", Colors.RED)
            sys.exit(1)

    cprint(f"\n📋 {len(scenarios)} senaryo yüklendi", Colors.CYAN)
    for s in scenarios:
        cprint(f"   {s['id']} — {s['query'][:60]}...", Colors.DIM)

    return scenarios


def load_ground_truth() -> dict:
    """Ground truth verilerini yükle."""
    if not GROUND_TRUTH_FILE.exists():
        cprint("ℹ️  ground_truth.json bulunamadı — sadece modeller-arası analiz yapılacak.", Colors.YELLOW)
        return {}

    gt = load_json(GROUND_TRUTH_FILE)
    gt.pop("_comment", None)

    # Kaç senaryo için dolu veri var?
    filled = sum(1 for v in gt.values() if v.get("feature_count") is not None)
    cprint(f"📊 Ground Truth: {filled}/{len(gt)} senaryo için ArcGIS referans verisi mevcut", Colors.CYAN)

    return gt


# ──────────────────────────────────────────────────────────────────────
# 3. n8n Köprüsü — Sıralı İstek
# ──────────────────────────────────────────────────────────────────────

async def run_scenario(
    client: httpx.AsyncClient,
    scenario: dict,
    index: int,
    total: int,
) -> dict:
    """
    Tek bir senaryoyu backend'e gönder, 7 modelin sonucunu al.
    """
    sid = scenario["id"]
    query = scenario["query"]

    cprint(
        f"\n{'─'*70}\n"
        f"  [{index}/{total}] {Colors.BOLD}{sid}{Colors.RESET} — {query}\n"
        f"{'─'*70}",
        "",
    )

    start_ts = time.time()
    try:
        resp = await client.post(
            ANALYZE_ENDPOINT,
            json={"message": query},
            timeout=REQUEST_TIMEOUT,
        )
        elapsed = time.time() - start_ts
        resp.raise_for_status()
        data = resp.json()

        # Her model için kısa durum bilgisi yazdır
        for m in ACTIVE_MODELS:
            key = m["key"]
            r = data.get(key, {})
            status = r.get("status", "missing")
            tool = r.get("selected_tool", "—")
            count = r.get("feature_count", 0)
            latency = r.get("processing_time_ms")
            latency_str = f"{latency/1000:.1f}s" if latency else "—"

            if status == "error":
                icon = "❌"
                color = Colors.RED
            elif status == "success":
                icon = "✅"
                color = Colors.GREEN
            elif status == "empty":
                icon = "⬜"
                color = Colors.YELLOW
            else:
                icon = "❓"
                color = Colors.DIM

            cprint(
                f"   {icon} {m['name']:16s} │ tool={tool:22s} │ "
                f"count={count:<4} │ latency={latency_str}",
                color,
            )

        cprint(f"   ⏱️  Toplam round-trip: {elapsed:.1f}s", Colors.DIM)
        return {"scenario_id": sid, "query": query, "results": data, "elapsed_s": elapsed, "error": None}

    except httpx.TimeoutException:
        cprint(f"   ❌ Zaman aşımı ({REQUEST_TIMEOUT}s)", Colors.RED)
        return {"scenario_id": sid, "query": query, "results": {}, "elapsed_s": time.time() - start_ts, "error": "timeout"}
    except Exception as exc:
        cprint(f"   ❌ Hata: {exc}", Colors.RED)
        return {"scenario_id": sid, "query": query, "results": {}, "elapsed_s": time.time() - start_ts, "error": str(exc)}


# ──────────────────────────────────────────────────────────────────────
# 4. GeoBenchX Puanlama (Uyarlanmış)
# ──────────────────────────────────────────────────────────────────────

def tool_matches(llm_tool: str, expected_tool: str) -> tuple:
    """
    LLM'in seçtiği araç, beklenen araçla eşleşiyor mu?
    Returns: (is_match: bool, is_partial: bool)
    
    GeoBenchX mantığı:
    - Tam eşleşme: tam isim veya alias eşleşmesi
    - Kısmi eşleşme: kelime bazlı overlap
    """
    if not llm_tool or not expected_tool:
        return False, False

    llm_norm = normalize_text(llm_tool)
    exp_norm = normalize_text(expected_tool)

    # Direkt eşleşme
    if llm_norm == exp_norm:
        return True, False

    # Alias eşleşmesi
    aliases = TOOL_ALIASES.get(expected_tool, [])
    alias_norms = [normalize_text(a) for a in aliases]
    if llm_norm in alias_norms:
        return True, False

    # LLM'in döndüğü tool, alias listesinin herhangi birini içeriyor mu?
    for alias in alias_norms:
        if alias in llm_norm or llm_norm in alias:
            return True, False

    # Kısmi eşleşme: kelime bazlı overlap
    llm_words = set(llm_norm.replace("_", " ").split())
    exp_words = set(exp_norm.replace("_", " ").split())
    for alias in alias_norms:
        exp_words.update(alias.replace("_", " ").split())

    overlap = llm_words & exp_words
    significant_overlap = {w for w in overlap if len(w) > 2}
    if significant_overlap:
        return False, True

    return False, False


def count_score_geobenchx(llm_count: int, ref_count: int) -> float:
    """
    GeoBenchX tarzı nesne sayısı puanlama (0–2 ölçeğinde).
    """
    if ref_count is None or ref_count <= 0:
        return 1.0  # Referans yok → nötr puan

    if llm_count == ref_count:
        return 2.0

    ratio = abs(llm_count - ref_count) / ref_count
    if ratio <= 0.05:
        return 1.8
    elif ratio <= 0.10:
        return 1.5
    elif ratio <= 0.20:
        return 1.0
    elif ratio <= 0.35:
        return 0.7
    elif ratio <= 0.50:
        return 0.4
    else:
        return 0.0


def spatial_compare(llm_geojson: dict, gt_names: list) -> dict:
    """
    LLM'in GeoJSON çıktısındaki feature isimlerini
    ArcGIS ground truth listesiyle karşılaştır.
    Precision / Recall / F1 hesapla.
    """
    if not gt_names:
        return {"precision": None, "recall": None, "f1": None, "tp": 0, "fp": 0, "fn": 0}

    # LLM feature isimlerini çıkar
    llm_names = set()
    for feature in llm_geojson.get("features", []):
        props = feature.get("properties", {})
        name = props.get("name", "") or props.get("Name", "") or ""
        if name:
            llm_names.add(normalize_text(name))

    gt_norm = set(normalize_text(n) for n in gt_names if n)

    if not gt_norm:
        return {"precision": None, "recall": None, "f1": None, "tp": 0, "fp": 0, "fn": 0}

    tp = len(llm_names & gt_norm)
    fp = len(llm_names - gt_norm)
    fn = len(gt_norm - llm_names)

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0

    return {"precision": round(precision, 4), "recall": round(recall, 4),
            "f1": round(f1, 4), "tp": tp, "fp": fp, "fn": fn}


def score_model_result(model_result: dict, scenario: dict, gt_entry: dict) -> dict:
    """
    Tek bir model sonucunu GeoBenchX metodolojisiyle puanla.
    
    Formül:
        Ground truth varsa:
            overall = 0.30 × tool + 0.35 × spatial_f1 + 0.20 × count + 0.15 × latency
        Ground truth yoksa:
            overall = 0.40 × tool + 0.60 × count_score_normalized
    """
    if not model_result or model_result.get("status") == "error":
        return {
            "tool_match": False,
            "tool_partial": False,
            "tool_score": 0.0,
            "count_score": 0.0,
            "spatial": {"precision": None, "recall": None, "f1": None, "tp": 0, "fp": 0, "fn": 0},
            "latency_score": 0.0,
            "overall_score": 0.0,
            "geobenchx_label": "❌ Yok",
            "has_error": True,
        }

    # 1. Tool Match
    llm_tool = model_result.get("selected_tool", "")
    expected_tool = scenario.get("expected_tool", "")
    is_match, is_partial = tool_matches(llm_tool, expected_tool)

    if is_match:
        tool_score = 2.0
    elif is_partial:
        tool_score = 1.0
    else:
        tool_score = 0.0

    # 2. Feature Count Score
    llm_count = model_result.get("feature_count", 0)
    if llm_count is None:
        # feature_count gelmemişse geojson'dan say
        geojson = model_result.get("geojson", {})
        llm_count = len(geojson.get("features", []))

    gt_count = gt_entry.get("feature_count") if gt_entry else None
    count_score = count_score_geobenchx(llm_count, gt_count)

    # 3. Spatial F1
    gt_names = gt_entry.get("feature_names", []) if gt_entry else []
    geojson = model_result.get("geojson", {"type": "FeatureCollection", "features": []})
    spatial = spatial_compare(geojson, gt_names)

    # 4. Latency Score (normalize: <2s → 2.0, 2-5s → 1.5, 5-10s → 1.0, >10s → 0.5)
    latency_ms = model_result.get("processing_time_ms")
    if latency_ms is not None:
        secs = latency_ms / 1000
        if secs < 2:
            latency_score = 2.0
        elif secs < 5:
            latency_score = 1.5
        elif secs < 10:
            latency_score = 1.0
        else:
            latency_score = 0.5
    else:
        latency_score = 1.0  # Bilinmiyor → nötr

    # 5. Overall Score
    has_gt = gt_entry and gt_entry.get("feature_count") is not None
    has_spatial = spatial["f1"] is not None

    if has_spatial:
        # Full formula: GT + spatial data mevcut
        spatial_f1_scaled = spatial["f1"] * 2.0  # 0–1 → 0–2 ölçeğine
        overall = (0.30 * tool_score +
                   0.35 * spatial_f1_scaled +
                   0.20 * count_score +
                   0.15 * latency_score)
    elif has_gt:
        # Count-based GT mevcut ama isim listesi yok
        overall = (0.40 * tool_score +
                   0.60 * count_score)
    else:
        # GT yok → sadece tool match + genel başarı
        success_bonus = 1.0 if model_result.get("status") == "success" and llm_count > 0 else 0.0
        overall = (0.40 * tool_score +
                   0.30 * success_bonus +
                   0.30 * latency_score)

    overall = round(min(2.0, overall), 2)

    # GeoBenchX etiketi
    if overall >= 1.8:
        label = "✅ Tam Eşleşme"
    elif overall >= 1.0:
        label = "⚡ Kısmi Eşleşme"
    elif overall > 0:
        label = "⚠️ Zayıf"
    else:
        label = "❌ Yok"

    return {
        "tool_match": is_match,
        "tool_partial": is_partial,
        "tool_score": round(tool_score, 2),
        "llm_tool": llm_tool,
        "llm_count": llm_count,
        "count_score": round(count_score, 2),
        "spatial": spatial,
        "latency_ms": latency_ms,
        "latency_score": round(latency_score, 2),
        "overall_score": overall,
        "geobenchx_label": label,
        "has_error": False,
    }


# ──────────────────────────────────────────────────────────────────────
# 5. Raporlama
# ──────────────────────────────────────────────────────────────────────

def print_summary_table(all_scores: list):
    """Konsola model × senaryo matris tablosu yazdır."""
    cprint(f"\n{'═'*90}", Colors.MAGENTA)
    cprint(f"  📊 BENCHMARK SONUÇ ÖZETİ — GeoBenchX Puanlama (0–2)", Colors.BOLD)
    cprint(f"{'═'*90}", Colors.MAGENTA)

    # Model ortalamaları
    model_avgs = {}
    model_tool_hits = {}
    for m in ACTIVE_MODELS:
        key = m["key"]
        scores = [
            entry["scores"][key]["overall_score"]
            for entry in all_scores
            if key in entry.get("scores", {}) and not entry["scores"][key].get("has_error")
        ]
        tool_hits = sum(
            1 for entry in all_scores
            if key in entry.get("scores", {}) and entry["scores"][key].get("tool_match")
        )
        total_scenarios = len(all_scores)
        model_avgs[key] = sum(scores) / len(scores) if scores else 0.0
        model_tool_hits[key] = f"{tool_hits}/{total_scenarios}"

    # Tablo başlığı
    header = f"  {'Model':16s} │ {'Ort.Puan':>8s} │ {'Tool Acc':>8s} │ {'Durum':14s}"
    cprint(f"\n{header}", Colors.BOLD)
    cprint(f"  {'─'*16}─┼─{'─'*8}─┼─{'─'*8}─┼─{'─'*14}", "")

    # Model satırları (puana göre sıralı)
    sorted_models = sorted(ACTIVE_MODELS, key=lambda m: model_avgs.get(m["key"], 0), reverse=True)
    medals = ["🥇", "🥈", "🥉"]

    for i, m in enumerate(sorted_models):
        key = m["key"]
        avg = model_avgs.get(key, 0)
        tool_acc = model_tool_hits.get(key, "0/0")
        medal = medals[i] if i < 3 else "  "

        if avg >= 1.5:
            color = Colors.GREEN
            status = "✅ Tam"
        elif avg >= 0.8:
            color = Colors.YELLOW
            status = "⚡ Kısmi"
        elif avg > 0:
            color = Colors.RED
            status = "⚠️ Zayıf"
        else:
            color = Colors.DIM
            status = "❌ Başarısız"

        cprint(
            f"  {medal}{m['name']:14s} │ {avg:>7.2f}/2 │ {tool_acc:>8s} │ {status}",
            color,
        )

    cprint(f"\n{'─'*90}", Colors.DIM)

    # Senaryo bazlı özet
    cprint(f"\n  📋 SENARYO DETAYLARI", Colors.BOLD)
    for entry in all_scores:
        sid = entry["scenario_id"]
        query = entry["query"][:45]
        cprint(f"\n  {Colors.CYAN}{sid}{Colors.RESET} — {query}")

        for m in ACTIVE_MODELS:
            key = m["key"]
            sc = entry["scores"].get(key, {})
            if sc.get("has_error"):
                cprint(f"     {m['name']:14s}  ❌ HATA", Colors.RED)
            else:
                label = sc.get("geobenchx_label", "—")
                overall = sc.get("overall_score", 0)
                tool = sc.get("llm_tool", "—") or "—"
                count = sc.get("llm_count", 0)
                cprint(
                    f"     {m['name']:14s}  {label:18s} │ score={overall:.2f} │ "
                    f"tool={tool:18s} │ count={count}",
                    "",
                )


def generate_json_report(all_scores: list, metadata: dict) -> Path:
    """JSON sonuç dosyası üret."""
    ts = timestamp_id()
    filepath = RESULTS_DIR / f"benchmark_{ts}.json"

    report = {
        "metadata": {
            **metadata,
            "timestamp": datetime.now().isoformat(),
            "models": [m["name"] for m in ACTIVE_MODELS],
            "scoring_method": "GeoBenchX adapted (0-2 scale)",
        },
        "results": all_scores,
        "model_summary": {},
    }

    # Model özetleri
    for m in ACTIVE_MODELS:
        key = m["key"]
        scores = [
            entry["scores"][key]["overall_score"]
            for entry in all_scores
            if key in entry["scores"] and not entry["scores"][key].get("has_error")
        ]
        errors = sum(
            1 for entry in all_scores
            if key in entry["scores"] and entry["scores"][key].get("has_error")
        )
        tool_hits = sum(
            1 for entry in all_scores
            if key in entry["scores"] and entry["scores"][key].get("tool_match")
        )

        report["model_summary"][key] = {
            "name": m["name"],
            "avg_score": round(sum(scores) / len(scores), 3) if scores else 0,
            "max_score": round(max(scores), 3) if scores else 0,
            "min_score": round(min(scores), 3) if scores else 0,
            "tool_accuracy": f"{tool_hits}/{len(all_scores)}",
            "error_count": errors,
            "scenarios_scored": len(scores),
        }

    save_json(report, filepath)
    cprint(f"\n💾 JSON rapor: {filepath}", Colors.GREEN)
    return filepath


def generate_markdown_report(all_scores: list, metadata: dict) -> Path:
    """Markdown rapor dosyası üret."""
    ts = timestamp_id()
    filepath = RESULTS_DIR / f"report_{ts}.md"

    lines = [
        "# CBS Agent Benchmark Raporu",
        f"\n> **Tarih:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"> **Metodoloji:** GeoBenchX (Krechetova & Kochedykov, 2025) — uyarlanmış",
        f"> **Puanlama:** 0–2 ölçek (2=Tam Eşleşme, 1=Kısmi, 0=Yok)",
        f"> **Senaryo Sayısı:** {len(all_scores)}",
        f"> **Model Sayısı:** {len(ACTIVE_MODELS)}",
        "",
        "## Model Karşılaştırma Özeti",
        "",
        "| # | Model | Ort. Puan | Tool Acc | Durum |",
        "|---|-------|-----------|----------|-------|",
    ]

    # Model sıralama
    model_data = []
    for m in ACTIVE_MODELS:
        key = m["key"]
        scores = [
            entry["scores"][key]["overall_score"]
            for entry in all_scores
            if key in entry["scores"] and not entry["scores"][key].get("has_error")
        ]
        tool_hits = sum(
            1 for entry in all_scores
            if key in entry["scores"] and entry["scores"][key].get("tool_match")
        )
        avg = sum(scores) / len(scores) if scores else 0
        model_data.append((m, avg, tool_hits))

    model_data.sort(key=lambda x: x[1], reverse=True)
    medals = ["🥇", "🥈", "🥉"]

    for i, (m, avg, tool_hits) in enumerate(model_data):
        medal = medals[i] if i < 3 else f"{i+1}."
        status = "✅" if avg >= 1.5 else ("⚡" if avg >= 0.8 else ("⚠️" if avg > 0 else "❌"))
        lines.append(
            f"| {medal} | {m['name']} | {avg:.2f}/2 | {tool_hits}/{len(all_scores)} | {status} |"
        )

    lines.extend(["", "## Senaryo Detayları", ""])

    for entry in all_scores:
        sid = entry["scenario_id"]
        query = entry["query"]
        lines.append(f"### {sid} — {query}")
        lines.append("")
        lines.append("| Model | Puan | Tool | Nesne | Durum |")
        lines.append("|-------|------|------|-------|-------|")

        for m in ACTIVE_MODELS:
            key = m["key"]
            sc = entry["scores"].get(key, {})
            if sc.get("has_error"):
                lines.append(f"| {m['name']} | 0.00 | — | — | ❌ Hata |")
            else:
                lines.append(
                    f"| {m['name']} | {sc.get('overall_score', 0):.2f} | "
                    f"{sc.get('llm_tool', '—')} | {sc.get('llm_count', 0)} | "
                    f"{sc.get('geobenchx_label', '—')} |"
                )
        lines.append("")

    lines.extend([
        "---",
        "",
        "*Bu rapor `benchmark/run_benchmark.py` tarafından otomatik üretilmiştir.*",
        f"*GeoBenchX: Krechetova & Kochedykov, GeoGenAgent '25, ACM SIGSPATIAL 2025*",
    ])

    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    cprint(f"📄 Markdown rapor: {filepath}", Colors.GREEN)
    return filepath


def update_excel(all_scores: list):
    """
    CBS_LLM_Dogruluk_ArcGIS_Karsilastirma.xlsx dosyasının
    '4. LLM_Cevaplari' ve '5. Puanlama' sayfalarını güncelle.
    """
    try:
        import openpyxl
    except ImportError:
        cprint("⚠️  openpyxl kurulu değil — Excel güncelleme atlanıyor.", Colors.YELLOW)
        return

    if not EXCEL_FILE.exists():
        cprint(f"⚠️  Excel dosyası bulunamadı: {EXCEL_FILE}", Colors.YELLOW)
        return

    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
    except Exception as exc:
        cprint(f"⚠️  Excel açılamadı: {exc}", Colors.YELLOW)
        return

    # Model sıralama (Excel'deki sıra: Claude, GPT-4o, Groq, Gemini)
    excel_models = ["claude", "gpt", "groq", "gemini"]

    # ── 4. LLM_Cevaplari sayfası ──
    sheet_name_4 = "4. LLM_Cevaplari"
    if sheet_name_4 in wb.sheetnames:
        ws4 = wb[sheet_name_4]
        # Veri satırları 5. satırdan başlar, her senaryo 4 model + 1 boş satır = 5 satır
        row = 5
        for entry in all_scores:
            for model_key in excel_models:
                sc = entry["scores"].get(model_key, {})
                if not sc.get("has_error", True):
                    # D: Seçilen Tool, E: LLM Sonuç Sayısı, G: Latency
                    ws4.cell(row=row, column=4, value=sc.get("llm_tool", ""))
                    ws4.cell(row=row, column=5, value=sc.get("llm_count", 0))
                    latency = sc.get("latency_ms")
                    if latency is not None:
                        ws4.cell(row=row, column=7, value=latency)
                row += 1
            row += 1  # Boş satır atlama

    # ── 5. Puanlama sayfası ──
    sheet_name_5 = "5. Puanlama"
    if sheet_name_5 in wb.sheetnames:
        ws5 = wb[sheet_name_5]
        row = 5
        for entry in all_scores:
            for model_key in excel_models:
                sc = entry["scores"].get(model_key, {})
                if not sc.get("has_error", True):
                    # D: Tool Match (1/0)
                    ws5.cell(row=row, column=4, value=1 if sc.get("tool_match") else 0)
                    # F: LLM Sonuç
                    ws5.cell(row=row, column=6, value=sc.get("llm_count", 0))
                row += 1
            row += 1

    try:
        wb.save(str(EXCEL_FILE))
        cprint(f"📊 Excel güncellendi: {EXCEL_FILE.name}", Colors.GREEN)
    except Exception as exc:
        cprint(f"⚠️  Excel kaydedilemedi: {exc}", Colors.YELLOW)


# ──────────────────────────────────────────────────────────────────────
# 6. Ana Akış
# ──────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(
        description="GeoBenchX Istanbul CBS Benchmark — n8n Köprüsü ile LLM Testi",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Örnekler:
  python benchmark/run_benchmark.py                  # Tam benchmark
  python benchmark/run_benchmark.py --dry-run        # Bağlantı olmadan test
  python benchmark/run_benchmark.py --scenario TR_01 # Tek senaryo
  python benchmark/run_benchmark.py --delay 5        # 5 sn ara ile
  python benchmark/run_benchmark.py --no-excel       # Excel güncelleme yapma
        """,
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Senaryo ve puanlama mantığını test et, gerçek istek atma")
    parser.add_argument("--scenario", type=str, default=None,
                        help="Tek senaryo çalıştır (ör: TR_01)")
    parser.add_argument("--delay", type=float, default=DEFAULT_DELAY,
                        help=f"Senaryolar arası bekleme süresi (saniye, varsayılan: {DEFAULT_DELAY})")
    parser.add_argument("--no-excel", action="store_true",
                        help="Excel dosyasını güncelleme")
    parser.add_argument("--api-url", type=str, default=None,
                        help="Backend API URL (varsayılan: http://localhost:8050)")

    args = parser.parse_args()

    if args.api_url:
        global API_BASE_URL, ANALYZE_ENDPOINT
        API_BASE_URL = args.api_url
        ANALYZE_ENDPOINT = f"{API_BASE_URL}/api/analyze"

    # Banner
    cprint(f"""
{'═'*70}
  📊 GeoBenchX Istanbul CBS Benchmark
  ─────────────────────────────────────
  Metodoloji : GeoBenchX (0–2) + Spatial F1
  Modeller   : {len(ACTIVE_MODELS)} LLM (Claude, GPT-4o, Groq, Gemini,
               DeepSeek, Qwen, GLM)
  Hedef      : {API_BASE_URL}
  Mod        : {'DRY-RUN (bağlantısız)' if args.dry_run else 'CANLI'}
{'═'*70}""", Colors.MAGENTA)

    # Senaryo ve GT yükle
    scenarios = load_scenarios(args.scenario)
    ground_truth = load_ground_truth()

    if args.dry_run:
        cprint("\n🧪 DRY-RUN — Sahte verilerle puanlama testi yapılıyor...\n", Colors.YELLOW)
        all_scores = []
        for scenario in scenarios:
            sid = scenario["id"]
            gt_entry = ground_truth.get(sid, {})

            # Sahte model sonuçları
            entry = {"scenario_id": sid, "query": scenario["query"], "scores": {}}
            for m in ACTIVE_MODELS:
                fake_result = {
                    "status": "success",
                    "selected_tool": scenario["expected_tool"],
                    "feature_count": 5,
                    "processing_time_ms": 2500,
                    "geojson": {"type": "FeatureCollection", "features": []},
                }
                entry["scores"][m["key"]] = score_model_result(fake_result, scenario, gt_entry)
            all_scores.append(entry)

        print_summary_table(all_scores)
        cprint("\n✅ Dry-run tamamlandı — puanlama mantığı çalışıyor.", Colors.GREEN)
        return

    # Health Check
    async with httpx.AsyncClient() as client:
        backend_ok = await health_check(client)
        if not backend_ok:
            cprint("\n❌ Backend'e bağlanılamıyor. Çıkılıyor.", Colors.RED)
            sys.exit(1)

        # n8n Probe
        n8n_ok = await n8n_probe(client)
        if not n8n_ok:
            cprint("\n⚠️  n8n webhook'ları yanıt vermiyor. Devam edilecek ama hatalar olabilir.", Colors.YELLOW)

        # Senaryoları sıralı çalıştır
        cprint(f"\n🚀 Benchmark başlıyor — {len(scenarios)} senaryo, {args.delay}s aralıkla\n", Colors.BOLD)
        all_results = []

        for i, scenario in enumerate(scenarios, 1):
            result = await run_scenario(client, scenario, i, len(scenarios))
            all_results.append(result)

            if i < len(scenarios):
                cprint(f"\n   ⏳ {args.delay}s bekleniyor...", Colors.DIM)
                await asyncio.sleep(args.delay)

    # Puanlama
    cprint(f"\n\n{'═'*70}", Colors.MAGENTA)
    cprint(f"  🧮 Puanlama hesaplanıyor...", Colors.BOLD)
    cprint(f"{'═'*70}", Colors.MAGENTA)

    all_scores = []
    for result in all_results:
        sid = result["scenario_id"]
        scenario = next(s for s in scenarios if s["id"] == sid)
        gt_entry = ground_truth.get(sid, {})

        entry = {
            "scenario_id": sid,
            "query": result["query"],
            "elapsed_s": result.get("elapsed_s"),
            "error": result.get("error"),
            "scores": {},
        }

        model_results = result.get("results", {})
        for m in ACTIVE_MODELS:
            key = m["key"]
            model_result = model_results.get(key, None)
            entry["scores"][key] = score_model_result(model_result, scenario, gt_entry)

        all_scores.append(entry)

    # Özet tablo
    print_summary_table(all_scores)

    # Raporlar
    metadata = {
        "scenario_count": len(scenarios),
        "model_count": len(ACTIVE_MODELS),
        "api_url": API_BASE_URL,
        "delay_s": args.delay,
    }

    generate_json_report(all_scores, metadata)
    generate_markdown_report(all_scores, metadata)

    if not args.no_excel:
        update_excel(all_scores)

    cprint(f"\n{'═'*70}", Colors.GREEN)
    cprint(f"  ✅ Benchmark tamamlandı!", Colors.BOLD)
    cprint(f"     Sonuçlar: {RESULTS_DIR}/", Colors.DIM)
    cprint(f"{'═'*70}\n", Colors.GREEN)


if __name__ == "__main__":
    asyncio.run(main())
